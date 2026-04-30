"""Read positions from the WM Growth Portfolio workbook.

The workbook is the user's ground truth (per `.memory/aaron.md`). The Holdings
table lives on the "Portfolio Overview" sheet and starts at row 7 (header).
The actual columns at the time of writing are::

    #, Ticker, Company Name, Sector, Weight %, Approx. Price,
    Analyst Rating, Avg. Price Target, Upside %, Key Catalyst, Risk / Watch

There is no "Target Weight" or "Cost Basis" column in this workbook — those
are sourced from ``config/targets.yaml`` and left as ``None`` respectively.
The tracker remains tolerant of either layout: if a column with one of the
expected aliases is present we use it; otherwise we fall back.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

DEFAULT_WORKBOOK = Path("/Users/aaronhart/Desktop/Claude Portfolio/WM_Growth_Portfolio_Apr2026.xlsx")
HOLDINGS_SHEET = "Portfolio Overview"
HEADER_ROW = 7

_COLUMN_ALIASES = {
    "ticker": {"ticker"},
    "company": {"company name", "company"},
    "sector": {"sector"},
    "current_weight": {"weight %", "current weight", "current weight %", "weight"},
    "target_weight": {"target weight", "target weight %", "target %"},
    "price": {"approx. price", "price", "last price"},
    "cost_basis": {"cost basis", "cost basis $", "avg cost", "cost"},
    "notes": {"notes", "key catalyst", "catalyst"},
    "risk": {"risk / watch", "risk", "watch"},
}


@dataclass
class Position:
    """One row of the Holdings sheet."""

    ticker: str
    company: str | None
    sector: str | None
    current_weight_pct: float  # percentage points (5.0 == 5.0%)
    target_weight_pct: float | None
    price: float | None
    cost_basis: float | None
    notes: str | None
    is_cash: bool = False
    is_etf: bool = False

    @property
    def drift_pct(self) -> float | None:
        """Current minus target, in percentage points."""
        if self.target_weight_pct is None:
            return None
        return self.current_weight_pct - self.target_weight_pct


@dataclass
class Portfolio:
    """All positions plus convenience aggregates."""

    positions: list[Position]
    asof: datetime
    source_path: Path
    sector_totals: dict[str, float] = field(default_factory=dict)
    cash_pct: float = 0.0

    @property
    def total_weight_pct(self) -> float:
        return sum(p.current_weight_pct for p in self.positions)

    @property
    def equity_positions(self) -> list[Position]:
        return [p for p in self.positions if not p.is_cash]

    def by_ticker(self, ticker: str) -> Position | None:
        for p in self.positions:
            if p.ticker == ticker.upper():
                return p
        return None


def _normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _build_column_index(header_row: tuple) -> dict[str, int]:
    index: dict[str, int] = {}
    for col_idx, raw in enumerate(header_row):
        norm = _normalize_header(raw)
        for key, aliases in _COLUMN_ALIASES.items():
            if norm in aliases and key not in index:
                index[key] = col_idx
    return index


def _to_pct(value: object) -> float | None:
    """Coerce a workbook weight cell to percentage points.

    The workbook stores 0.065 (=6.5%). If a value already exceeds 1.5 we
    assume the user typed it as percentage points. Round to 6dp so the
    decimal-to-percent multiplication doesn't surface floating-point cruft
    (e.g. 0.035 * 100 = 3.5000000000000004 -> falsely trips a 0.50% drift cap).
    """
    if value is None or value == "" or value == "—":
        return None
    try:
        f = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    out = f * 100.0 if abs(f) <= 1.5 else f
    return round(out, 6)


def _to_float(value: object) -> float | None:
    if value is None or value == "" or value == "—":
        return None
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _is_etf_sector(sector: str | None) -> bool:
    if not sector:
        return False
    return "etf" in sector.lower()


def load_targets(targets_path: Path) -> dict[str, float]:
    """Load `config/targets.yaml`. Returns ticker -> target percentage points."""
    if not targets_path.exists():
        return {}
    with targets_path.open() as fh:
        raw = yaml.safe_load(fh) or {}
    out: dict[str, float] = {}
    for k, v in (raw.get("targets") or {}).items():
        out[str(k).upper()] = float(v)
    return out


DEFAULT_POSITIONS_YAML = Path(__file__).resolve().parents[2] / "config" / "positions.yaml"


def _load_portfolio_from_yaml(path: Path) -> Portfolio:
    """Reconstitute a Portfolio from the committed yaml snapshot.

    Used as a fallback when the xlsx is missing (e.g. on Streamlit Cloud
    where the local file system has no access to the user's Desktop).
    """
    with path.open() as fh:
        raw = yaml.safe_load(fh) or {}
    # PyYAML auto-parses ISO timestamps into ``datetime`` rather than ``str``,
    # so accept both shapes here.
    asof_raw = raw.get("asof_iso")
    if isinstance(asof_raw, datetime):
        asof = asof_raw
    elif isinstance(asof_raw, str):
        try:
            asof = datetime.fromisoformat(asof_raw)
        except ValueError:
            asof = datetime.fromtimestamp(path.stat().st_mtime)
    else:
        asof = datetime.fromtimestamp(path.stat().st_mtime)
    positions: list[Position] = []
    cash_pct = 0.0
    for row in raw.get("positions") or []:
        pos = Position(
            ticker=str(row["ticker"]).upper(),
            company=row.get("company"),
            sector=row.get("sector"),
            current_weight_pct=float(row["current_weight_pct"]),
            target_weight_pct=(
                float(row["target_weight_pct"]) if row.get("target_weight_pct") is not None else None
            ),
            price=float(row["price"]) if row.get("price") is not None else None,
            cost_basis=float(row["cost_basis"]) if row.get("cost_basis") is not None else None,
            notes=row.get("notes"),
            is_cash=bool(row.get("is_cash", False)),
            is_etf=bool(row.get("is_etf", False)),
        )
        positions.append(pos)
        if pos.is_cash:
            cash_pct = pos.current_weight_pct
    sector_totals: dict[str, float] = {}
    for p in positions:
        if p.is_cash or not p.sector:
            continue
        head = p.sector.split("–")[0].split("-")[0].strip()
        sector_totals[head] = sector_totals.get(head, 0.0) + p.current_weight_pct
    return Portfolio(
        positions=positions,
        asof=asof,
        source_path=path,
        sector_totals=sector_totals,
        cash_pct=cash_pct,
    )


def load_portfolio(
    workbook_path: Path = DEFAULT_WORKBOOK,
    targets_path: Path | None = None,
    positions_yaml: Path = DEFAULT_POSITIONS_YAML,
) -> Portfolio:
    """Read the Holdings sheet and return a populated ``Portfolio``.

    Resolution order:
    1. If ``workbook_path`` exists, read the xlsx directly (live, authoritative).
    2. Else if ``positions_yaml`` exists, fall back to the committed snapshot
       (used on Streamlit Cloud where the user's xlsx isn't reachable).
    3. Else raise.
    """
    if not workbook_path.exists():
        if positions_yaml.exists():
            return _load_portfolio_from_yaml(positions_yaml)
        raise FileNotFoundError(
            f"Neither workbook ({workbook_path}) nor positions yaml "
            f"({positions_yaml}) found."
        )
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if HOLDINGS_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{HOLDINGS_SHEET}' missing in {workbook_path}")
    ws = wb[HOLDINGS_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < HEADER_ROW:
        raise ValueError(f"Sheet has only {len(rows)} rows; expected header at row {HEADER_ROW}")
    header = rows[HEADER_ROW - 1]
    cols = _build_column_index(header)
    if "ticker" not in cols or "current_weight" not in cols:
        raise ValueError(
            f"Workbook header missing required columns. Saw: {[_normalize_header(h) for h in header]}"
        )

    targets = load_targets(targets_path) if targets_path else {}

    positions: list[Position] = []
    cash_pct = 0.0
    for row in rows[HEADER_ROW:]:
        ticker_raw = row[cols["ticker"]]
        if ticker_raw is None:
            continue
        ticker = str(ticker_raw).strip().upper()
        if not ticker or ticker in {"TOTAL"}:
            continue

        company = row[cols["company"]] if "company" in cols else None
        sector = row[cols["sector"]] if "sector" in cols else None
        weight = _to_pct(row[cols["current_weight"]])
        if weight is None:
            continue
        target = _to_pct(row[cols["target_weight"]]) if "target_weight" in cols else None
        if target is None:
            target = targets.get(ticker)
        price = _to_float(row[cols["price"]]) if "price" in cols else None
        cost_basis = _to_float(row[cols["cost_basis"]]) if "cost_basis" in cols else None
        notes = row[cols["notes"]] if "notes" in cols else None

        is_cash = ticker == "CASH"
        is_etf = _is_etf_sector(str(sector) if sector else None)
        pos = Position(
            ticker=ticker,
            company=str(company) if company else None,
            sector=str(sector) if sector else None,
            current_weight_pct=weight,
            target_weight_pct=target,
            price=price,
            cost_basis=cost_basis,
            notes=str(notes) if notes else None,
            is_cash=is_cash,
            is_etf=is_etf,
        )
        positions.append(pos)
        if is_cash:
            cash_pct = weight

    sector_totals: dict[str, float] = {}
    for p in positions:
        if p.is_cash or not p.sector:
            continue
        # Roll sub-sectors ("Technology - AI / Cloud") up to their parent.
        head = p.sector.split("–")[0].split("-")[0].strip()
        sector_totals[head] = sector_totals.get(head, 0.0) + p.current_weight_pct

    asof = datetime.fromtimestamp(workbook_path.stat().st_mtime)
    return Portfolio(
        positions=positions,
        asof=asof,
        source_path=workbook_path,
        sector_totals=sector_totals,
        cash_pct=cash_pct,
    )
